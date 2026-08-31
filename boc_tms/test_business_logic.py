"""Standalone unit tests for database.py business logic (not shipped to user)."""
import os
import datetime
import pandas as pd
import database
import excel_sync

# Use a throwaway DB so this doesn't disturb seeded demo data
database.DB_PATH = "test_boc.db"
excel_sync.BACKUP_PATH = "test_backup.xlsx"
for f in (database.DB_PATH, excel_sync.BACKUP_PATH):
    if os.path.exists(f):
        os.remove(f)

database.init_db()
database.seed_if_empty()

drivers = database.get_drivers()
vehicles = database.get_vehicles()
departments = database.get_departments()
d1, d2 = int(drivers.iloc[0]["id"]), int(drivers.iloc[1]["id"])
v1, v2 = int(vehicles.iloc[0]["id"]), int(vehicles.iloc[1]["id"])
dep = int(departments.iloc[0]["id"])

results = []

# 1) Baseline booking should succeed
ok, conflicts = database.add_appointment("2026-09-01", "09:00", "11:00", d1, v1, dep, "Test A", "tester")
results.append(("Baseline booking succeeds", ok and not conflicts))

# 2) Overlapping time, SAME driver, different vehicle -> must be rejected
ok, conflicts = database.add_appointment("2026-09-01", "10:00", "12:00", d1, v2, dep, "Test B", "tester")
results.append(("Driver double-booking rejected", not ok and len(conflicts) > 0))

# 3) Overlapping time, SAME vehicle, different driver -> must be rejected
ok, conflicts = database.add_appointment("2026-09-01", "10:00", "12:00", d2, v1, dep, "Test C", "tester")
results.append(("Vehicle double-booking rejected", not ok and len(conflicts) > 0))

# 4) Non-overlapping time, same driver+vehicle -> must succeed
ok, conflicts = database.add_appointment("2026-09-01", "11:00", "12:00", d1, v1, dep, "Test D", "tester")
results.append(("Back-to-back non-overlapping booking succeeds", ok and not conflicts))

# 5) Different date, same driver+vehicle+time -> must succeed
ok, conflicts = database.add_appointment("2026-09-02", "09:00", "11:00", d1, v1, dep, "Test E", "tester")
results.append(("Same slot different date succeeds", ok and not conflicts))

# 6) Cancelling an appointment should free the slot for a new booking
appts = database.get_appointments("2026-09-01", "2026-09-01")
first_id = int(appts.iloc[0]["id"])
database.update_appointment_status(first_id, "Cancelled")
ok, conflicts = database.add_appointment("2026-09-01", "09:00", "11:00", d1, v1, dep, "Test F", "tester")
results.append(("Cancelled appointment frees the slot", ok and not conflicts))

# 7) Driver/vehicle status updates persist
database.update_driver_status(d1, "On Leave")
d1_status = database.get_drivers()
d1_status = d1_status[d1_status["id"] == d1].iloc[0]["status"]
results.append(("Driver status update persists", d1_status == "On Leave"))

# 8) Add driver / add vehicle
database.add_driver("Test Driver", "B9999999", "0770000000", "Test Depot", lat=6.9, lon=79.8)
results.append(("New driver appears in list", "Test Driver" in database.get_drivers()["name"].values))

database.add_vehicle("TEST-0001", "Van", 8, lat=6.9, lon=79.8)
results.append(("New vehicle appears in list", "TEST-0001" in database.get_vehicles()["plate_no"].values))

# 9) Excel backup file gets created and contains expected sheets
excel_sync.force_sync_blocking()
import openpyxl
wb = openpyxl.load_workbook(excel_sync.BACKUP_PATH)
results.append(("Excel backup has all sheets",
                 set(wb.sheetnames) == {"Summary", "Drivers", "Vehicles", "Departments",
                                        "Appointments", "Trip Requests", "Leave Requests"}))
results.append(("Excel backup Drivers sheet row count matches DB",
                 wb["Drivers"].max_row - 1 == len(database.get_drivers())))

# 10) Departments management + linked login creation
new_dept_id = database.add_department("Test Branch", "Test City", lat=7.0, lon=80.0)
results.append(("New department appears in list", "Test Branch" in database.get_departments()["name"].values))

database.create_login("test_dept_user", "TestPass123", "Test Branch", "Department", linked_department_id=new_dept_id)
login_check = database.verify_login("test_dept_user", "TestPass123")
results.append(("New department login authenticates", login_check is not None and login_check["role"] == "Department"))
results.append(("username_exists detects existing username", database.username_exists("test_dept_user")))
results.append(("username_exists returns False for unused name", not database.username_exists("nobody_here")))

# 11) Department request -> approve/reject/cancel workflow
database.add_trip_request(new_dept_id, "2026-10-01", "09:00", "10:00", "Test request", "test_dept_user")
pend = database.get_trip_requests(status="Pending", department_id=new_dept_id)
results.append(("Trip request created as Pending", len(pend) == 1))

req_id = int(pend.iloc[0]["id"])
ok, conflicts = database.approve_trip_request(req_id, d2, v2)
results.append(("Approving a request creates an appointment", ok and not conflicts))
approved = database.get_trip_requests(status="Approved", department_id=new_dept_id)
results.append(("Approved request links to a new appointment", not approved.empty and pd.notna(approved.iloc[0]["appointment_id"])))

database.add_trip_request(new_dept_id, "2026-10-02", "09:00", "10:00", "Test cancel", "test_dept_user")
pend2 = database.get_trip_requests(status="Pending", department_id=new_dept_id)
cancel_id = int(pend2.iloc[0]["id"])
database.cancel_trip_request(cancel_id)
cancelled = database.get_trip_requests(status="Cancelled", department_id=new_dept_id)
results.append(("Cancelling a pending request works", len(cancelled) == 1))

# 12) Vehicle compliance tracking
test_vehicle = database.get_vehicles().iloc[0]
overdue_date = (datetime.date.today() - datetime.timedelta(days=5)).isoformat()
healthy_date = (datetime.date.today() + datetime.timedelta(days=365)).isoformat()
database.update_vehicle_compliance(int(test_vehicle["id"]), insurance_expiry=overdue_date,
                                    revenue_license_expiry=healthy_date, next_service_due=healthy_date)
attention = database.vehicles_needing_attention(warn_days=30)
flagged = attention[attention["id"] == test_vehicle["id"]]
results.append(("Overdue insurance flags the vehicle in vehicles_needing_attention",
                 not flagged.empty and "OVERDUE" in flagged.iloc[0]["issues"]))

database.update_vehicle_compliance(int(test_vehicle["id"]), insurance_expiry=healthy_date,
                                    revenue_license_expiry=healthy_date, next_service_due=healthy_date)
attention2 = database.vehicles_needing_attention(warn_days=30)
results.append(("Updating compliance dates clears the alert",
                 test_vehicle["id"] not in attention2["id"].values if not attention2.empty else True))

# 13) Driver leave request workflow
test_driver_id = int(database.get_drivers().iloc[0]["id"])
today_iso = datetime.date.today().isoformat()
end_iso = (datetime.date.today() + datetime.timedelta(days=2)).isoformat()
database.add_leave_request(test_driver_id, today_iso, end_iso, "Test leave")
leave_pending = database.get_leave_requests(status="Pending", driver_id=test_driver_id)
results.append(("Leave request created as Pending", len(leave_pending) == 1))

leave_id = int(leave_pending.iloc[0]["id"])
database.approve_leave_request(leave_id)
leave_approved = database.get_leave_requests(status="Approved", driver_id=test_driver_id)
results.append(("Approving a leave request marks it Approved", len(leave_approved) == 1))
driver_status_after = database.get_drivers()
driver_status_after = driver_status_after[driver_status_after["id"] == test_driver_id].iloc[0]["status"]
results.append(("Approved leave covering today sets driver status to On Leave",
                 driver_status_after == "On Leave"))

database.add_leave_request(test_driver_id, "2026-12-01", "2026-12-03", "Test reject leave")
leave_pending2 = database.get_leave_requests(status="Pending", driver_id=test_driver_id)
leave_id2 = int(leave_pending2.iloc[0]["id"])
database.reject_leave_request(leave_id2, "Coverage unavailable")
leave_rejected = database.get_leave_requests(status="Rejected", driver_id=test_driver_id)
results.append(("Rejecting a leave request stores the reason",
                 len(leave_rejected) == 1 and leave_rejected.iloc[0]["decision_note"] == "Coverage unavailable"))

# 14) Schema migration: an old-style vehicles table (no compliance columns) upgrades cleanly
import sqlite3
migrate_test_db = "test_migration.db"
if os.path.exists(migrate_test_db):
    os.remove(migrate_test_db)
old_conn = sqlite3.connect(migrate_test_db)
old_conn.execute("""CREATE TABLE vehicles (
    id INTEGER PRIMARY KEY AUTOINCREMENT, plate_no TEXT UNIQUE NOT NULL,
    vehicle_type TEXT NOT NULL, capacity INTEGER,
    status TEXT NOT NULL DEFAULT 'Available', lat REAL, lon REAL)""")
old_conn.execute("INSERT INTO vehicles (plate_no, vehicle_type, capacity, status, lat, lon) "
                  "VALUES ('MIGRATE-01','Van',8,'Available',6.9,79.8)")
old_conn.commit()
old_conn.close()

original_db_path = database.DB_PATH
database.DB_PATH = migrate_test_db
database.init_db()
migrated = database.get_vehicles()
results.append(("Schema migration adds compliance columns without losing data",
                 "insurance_expiry" in migrated.columns and
                 migrated.iloc[0]["plate_no"] == "MIGRATE-01"))
database.DB_PATH = original_db_path
os.remove(migrate_test_db)

# --- report ---
print("\n=== BUSINESS LOGIC TEST RESULTS ===")
all_pass = True
for name, passed in results:
    print(f"{'PASS' if passed else 'FAIL'} - {name}")
    all_pass = all_pass and passed
print("\nALL PASS:", all_pass)

# cleanup
for f in (database.DB_PATH, excel_sync.BACKUP_PATH):
    if os.path.exists(f):
        os.remove(f)